import openpyxl,os
import datetime
import matplotlib.pyplot as plt
os.chdir(r'd:\!scrap')
wb = openpyxl.load_workbook('Logitech G331-Leatherette.xlsx')

sheet = wb['Sheet1']
price=[]
day=[]
columnA = 1
columnB = 2
row = 2
while(sheet.cell(row,columnB).value!=None):
    price.append(int(sheet.cell(row,columnA).value))
    day.append(str(sheet.cell(row,columnB).value))
    row += 1

print(price)
print(day)

plt.title('Logitech G331 Leatherette Trend:')
plt.plot(day,price,'-o' ,color='red')
plt.legend(['Price on Flipkart'] , loc = "upper right" )
plt.xlabel('DAY')
plt.ylabel('PRICE')
plt.tick_params(axis='x', which='major', labelsize=8)
plt.tight_layout()
plt.show()

